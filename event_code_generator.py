
requirements = ["openpyxl"]

def check_requirement(package):
    try:
        exec("import {0}".format(package))
    except ModuleNotFoundError:
        inquiry = input("This script requires {0}. Do you want to install {0}? [y/n]".format(package))
        while (inquiry != "y") and (inquiry != "n"):
            inquiry = input("This script requires {0}. Do you want to install {0}? [y/n]".format(package))
        if inquiry == "y":
            import os
            print("Execute commands: pip3 install {0}".format(package))
            os.system("pip3 install {0}".format(package))
        else:
            print("{0} is missing, so the program exits!".format(package))
            exit(-1)

for requirement in requirements:
    check_requirement(requirement)

import os
import time
from openpyxl import load_workbook
import argparse

indentation = "    "
super_class = "NSObject"
analytic_prorocol = "AnalyticEvent"
class_prefix = "TPP"
excel_file_path = None
analytic_manager_file_path = None
generate_file_path = class_prefix + "GenerateAnalytics.swift"
default_analytics_file_name = "AnalyticsManager.m"

class Column():
    EVENT_NAME = 0
    ATTRIBUTE_NAME   = 1
    ATTRIBUTE_TYPE = 2
    ATTRIBUTE_DESC = 3
    ATTRIBUTE_VALUE = 4


class EventKey():
    NAME = "event_name"
    ATTRIBUTES = "attributes"


class AttributeType():
    STRING = "String"
    NUMBER = "Number"


class Event():
    class_name = ""
    def __init__(self, name, attributes):
        self.name = name
        self.attributes = attributes    

    def generateCode(self):
        self.class_name = class_prefix + toVariableString(self.name) + "Event"

        lines = ["@objc class " + self.class_name + ": " + super_class + ", " + analytic_prorocol + " {\n"]
        lines.append(indentation + 'static var name: String = "' + self.name + '"\n')
        lines.append(indentation + 'var parametrs = [AnyHashable : Any]()\n')

        for attr in self.attributes:
            attr_lines = attr.generateCode()
            if attr_lines is not None:
                lines.extend(map(lambda x: indentation + x, attr_lines))

        init_func_code = self.generate_init_func()
        if init_func_code is not None:
            lines.extend(map(lambda x: indentation + x, init_func_code))

        lines.append("}\n")
        return lines

    def generate_init_func(self):
        if len(self.attributes) > 0:
            parameter_lines = []
            first_line = "init("
            for att in self.attributes:
                p_n = toVariableString(att.name, first_lower=True)
                first_line += (p_n + ": " + att.typeName() + " ")
                if att.enum_attribute:
                    parameter_lines.append(indentation + 'parametrs["' + att.name + '"] = ' + p_n + '.rawValue\n')
                else:
                    parameter_lines.append(indentation + 'parametrs["' + att.name + '"] = ' + p_n + '\n')
            first_line = first_line.strip()
            first_line += ") {\n"
            return [first_line] + parameter_lines + ["}\n"]
                

class Attribute():
    values = []
    enum_attribute = False
    def __init__(self, name, type=None, values=[]):
        self.name = name
        self.type = type
        self.values = values

    def appendValue(self, value):
        self.values.append(value)

    def generateCode(self):
        if len(self.values) > 1 and self.type == AttributeType.STRING:
            enum_attribute = True
            enum_name = toVariableString(self.name)
            lines = ["enum " + enum_name + ": " + "String {\n"]
            for v in self.values:
                lines.append(indentation + "case " + toVariableString(v, first_lower=True) + " = " + '"' + v + '"\n')
            lines.append("}\n")
            return lines

    def typeName(self):
        if self.type == AttributeType.STRING:
            if len(self.values) > 0:
                return toVariableString(self.name)
            return "String"

        if self.type == AttributeType.NUMBER:
            return "Int"
        return ""
            

def toVariableString(string, first_lower=False):
    tripedStr = str(string).strip()
    str1 = "".join(map(lambda x:x.capitalize(), tripedStr.split(" ")))
    str2 = "".join(map(lambda x:x[:1].upper() + x[1:], str1.split("_")))
    if first_lower:
        return str2[:1].lower() + str2[1:]
    return str2


def get_current_time():
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())


def parseEventExcel(filename):
    work_book = load_workbook(filename=filename)
    sheet = work_book["Sheet1"]

    event_name = None
    attr_name = None
    attributes = dict()

    event_list = []
    for row in sheet.iter_rows(min_row=2):
        event_name_col = row[Column.EVENT_NAME].value
        attr_value = row[Column.ATTRIBUTE_VALUE].value
        attr_type = row[Column.ATTRIBUTE_TYPE].value
        if event_name_col is not None and event_name_col != event_name:
            # 新的事件
            attr_name = row[Column.ATTRIBUTE_NAME].value
            if event_name is not None:
                event_list.append(Event(name=event_name, attributes=list(attributes.values())))
            event_name = event_name_col
            if attr_name is not None:
                attributes = {attr_name: Attribute(name=attr_name, type=attr_type, values=[attr_value])}
            else:
                attributes = {}

        elif event_name is not None:
            attr_name = row[Column.ATTRIBUTE_NAME].value or attr_name
            # 与上一行为同一事件
            attribute = attributes.get(attr_name, None)
            if attribute is None:
                attribute = Attribute(name=attr_name, type=attr_type)
                attributes[attr_name] = attribute
            if attr_value is not None:
                attribute.appendValue(attr_value)
        else:
            raise "event_name error"
    event_list.append(Event(name=event_name, attributes=list(attributes.values())))
    print("Successfilly parsed excel and got {0} events".format(len(event_list)))
    return event_list

def generate_code(event_list, filename):
    path, name = os.path.split(filename)
    code_list = ["// \n", "// {0}\n".format(name), "// Generated by event_code_generator on {0}\n".format(get_current_time()), "// \n", "\n", "\n", "import Foundation\n", "\n"]
    for event in event_list:
        event_code = event.generateCode ()
        if event_code is not None:
            code_list.extend(event_code)
    
    with open(filename, 'w') as f:
        f.writelines(code_list)

    print("Successfully generated code and saved into file " + filename)


def registerEvents(events, analytics_file_path):
    new_lines = []
    line_prefix_space = ""
    with open(analytics_file_path, 'r') as f:
        lines = f.readlines()
        count = 1
        events_map_begin = False
        events_end_index = 1
        for line in lines:
            new_lines.append(line)
            if line.strip().startswith("self.eventsMap = @{"):
                events_map_begin = True
            if events_map_begin and line.strip().startswith("}"):
                line_prefix_space = line.split("}")[0]
                events_end_index = count
                events_map_begin = False
            count += 1
    for event in events:
        new_lines.insert(events_end_index - 1, line_prefix_space + event.class_name + ".name: @[mixpanel],\n")
    current_time = get_current_time()
    new_lines.insert(events_end_index - 1, line_prefix_space + "// generate by script at " + current_time + "\n")
    with open(analytics_file_path, 'w+') as f:
        f.writelines(new_lines)
    
    print("Successfully inserted events into file {0}, line {1}".format(analytics_file_path, events_end_index))


def find_analytics_file(start, name):
    for relpath, dirs, files in os.walk(start):
        if name in files:
            full_path = os.path.join(start, relpath, name)
            path = os.path.normpath(os.path.abspath(full_path))
            inquiry = input("Finde file at {0}. Do you want to use this file? [y/n]".format(path))
            while (inquiry != "y") and (inquiry != "n"):
                inquiry = input("Finde file at {0}. Do you want to use this file? [y/n]".format(path))
            if inquiry.lower() == 'y':
                return path


if __name__ == '__main__':
    
    parser = argparse.ArgumentParser(description='Generate event excel to swift event class code and register to Analyticsmanager.m')
    parser.add_argument('--event_excel', type=str,help='File path for excel')
    parser.add_argument('--code_file', type=str,help='File path for generated code')
    parser.add_argument('--analytics', type=str,help='File path for AnalyticsManager.m')
    args = parser.parse_args()
    
    if args.event_excel is None:
        raise Exception("Need event_excel argument for parse")
    excel_file_path = args.event_excel
    print("event_excel:" + args.event_excel)
    if args.code_file is None:
        print("'code_file' argument is missing, generated code will save at " + os.path.abspath(os.path.dirname(__file__)) + "/" + generate_file_path)
    generate_file_path = args.code_file or generate_file_path
    if args.analytics is None:
        print("'ananlytics' argument is missing, The file AnalyticsManager will be looking for locally...")
        analytic_manager_file_path = find_analytics_file(os.path.abspath(os.path.dirname(__file__)), "AnalyticsManager.m")
        if analytic_manager_file_path is None:
            raise Exception("AnalyticsManager.m not found.")
    else:
        analytic_manager_file_path = args.analytics

    event_list = parseEventExcel(filename=excel_file_path)
    generate_code(event_list=event_list, filename=generate_file_path)
    registerEvents(event_list, analytics_file_path=analytic_manager_file_path)

    print("Generate successfully!!!")

